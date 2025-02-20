# src/services/parser.py
import openpyxl

# Определяем диапазоны строк для дней (например, Пн: строки 5–13, Вт: 14–22 и т.д.)
DAY_RANGES = {
    0: range(5, 14),   # Понедельник
    1: range(14, 23),  # Вторник
    2: range(23, 32),  # Среда
    3: range(32, 41),  # Четверг
    4: range(41, 50),  # Пятница
    5: range(50, 59),  # Суббота
}

def process_lesson_cell(cell):
    """
    Обрабатывает ячейку урока и возвращает список записей.
    Каждая запись – словарь с ключами:
      - text: текст урока,
      - program: "ФГОС" или "МП", если встречается,
      - is_language: True, если в тексте есть слово "язык",
      - cell_color: "blue" или "green" (определяется по RGB).
    """
    text = cell.value
    if text is None:
        return []
    text = str(text).strip()

    # Определяем цвет ячейки
    color_rgb = cell.fill.fgColor.rgb
    cell_color = "default"
    if color_rgb:
        if color_rgb.startswith("FF"):
            color_hex = color_rgb[2:]
        else:
            color_hex = color_rgb
        if color_hex.upper() == "7DB4F0":  # синий
            cell_color = "blue"
        elif color_hex.upper() == "70AD47":  # зелёный
            cell_color = "green"
        else:
            cell_color = "default"

    # Если вариантов несколько, разделённых символом "|"
    parts = text.split("|")
    entries = []
    for part in parts:
        part = part.strip()
        program_marker = None
        if "ФГОС" in part:
            program_marker = "ФГОС"
        elif "МП" in part:
            program_marker = "МП"
        is_language = "язык" in part.lower()
        entries.append({
            "text": part,
            "program": program_marker,
            "is_language": is_language,
            "cell_color": cell_color
        })
    return entries

def parse_schedule(excel_path: str):
    print(f"[parser.py] Открываем файл: {excel_path}")
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # Считываем названия групп из 4-й строки (E4..S4).
    group_names = {}
    last_value = None
    for col_idx in range(5, 20):  # E=5, ..., S=19 (всего может быть 15+ групп)
        cell_value = sheet.cell(row=4, column=col_idx).value
        if cell_value:
            last_value = str(cell_value).strip()
            group_names[col_idx] = last_value
        else:
            group_names[col_idx] = last_value  # Если пустая, используем предыдущее значение.
        print(f"DEBUG: col={col_idx} -> group_name='{group_names[col_idx]}'")

    # Инициализируем структуру: schedule_data[group][day] = список уроков.
    schedule_data = {}
    for col_idx, group in group_names.items():
        if group is None:
            continue
        if group not in schedule_data:
            schedule_data[group] = {day: [] for day in DAY_RANGES}

    for day, rows in DAY_RANGES.items():
        for row in rows:
            # Читаем номер пары из столбца C (column 3)
            lesson_number = sheet.cell(row=row, column=3).value
            # Читаем время из столбца D (column 4)
            time_cell = sheet.cell(row=row, column=4)
            time_val = time_cell.value
            if not time_val:
                continue
            time_text = str(time_val).strip()
            for col_idx, group in group_names.items():
                if group is None:
                    continue
                cell = sheet.cell(row=row, column=col_idx)
                entries = process_lesson_cell(cell)
                if entries:
                    schedule_data[group][day].append({
                        "number": lesson_number,
                        "time": time_text,
                        "entries": entries
                    })
    print("\n[parser.py] Итоговое расписание (schedule_data):")
    for grp, days in schedule_data.items():
        print(f"Группа '{grp}':")
        for d, lessons in days.items():
            print(f"  День {d} => {lessons}")
    print("[parser.py] Конец вывода расписания.\n")
    return schedule_data
